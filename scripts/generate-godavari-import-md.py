from __future__ import annotations

import ast
import calendar
import datetime as dt
import re
from pathlib import Path
from typing import Any

import openpyxl
from openpyxl.utils import get_column_letter


ROOT = Path(__file__).resolve().parents[1]
WORKBOOK = ROOT / "Dataset" / "GPS-Maintenance FY 2019 to 2026.xlsx"
OUTPUT = ROOT / "docs" / "data-import" / "godavari-puram-import.sql.md"

SOCIETY_ID = "soc_godavari_puram"
ADMIN_USER_ID = "user_gps_chairman"
ADMIN_MEMBERSHIP_ID = "membership_gps_chairman"
SOCIETY_NAME = "GODAVARI PURAM CO-OPERATIVE HOUSING SOCIETY LTD"
JOIN_CODE = "GODAVA2026"
TEMP_PASSWORD = "Gps@2026"
PASSWORD_HASH = "$2b$12$nJ9vaiJF8FF1/tuV/C5vneXlec3TNKSQzgKZX/hKMdgk9ML.Q6qyW"
GENERATED_DOMAIN = "godavaripuram.local"

MONTH_NUMBERS = {
    "JAN": 1,
    "FEB": 2,
    "MAR": 3,
    "APR": 4,
    "APRIL": 4,
    "MAY": 5,
    "JUN": 6,
    "JUNE": 6,
    "JUL": 7,
    "JULY": 7,
    "AUG": 8,
    "SEP": 9,
    "SEPT": 9,
    "OCT": 10,
    "NOV": 11,
    "DEC": 12,
}


def sql(value: Any) -> str:
    if value is None:
        return "NULL"
    if isinstance(value, bool):
        return "TRUE" if value else "FALSE"
    if isinstance(value, (int, float)):
        return str(int(value)) if float(value).is_integer() else repr(float(value))
    text = str(value)
    return "'" + text.replace("'", "''") + "'"


def clean_phone(value: Any) -> str | None:
    if value is None:
        return None
    digits = re.sub(r"\D", "", str(value))
    if digits.endswith("0") and len(digits) == 11:
        digits = digits[:-1]
    return digits or None


def clean_person_name(value: Any) -> str:
    name = re.sub(r"\s+", " ", str(value or "")).strip()
    honorific_pattern = re.compile(
        r"^(?:shri|shree|sau|smt|mr|mrs|ms|miss|dr)(?:\.\s*|\s+)",
        re.IGNORECASE,
    )
    while honorific_pattern.match(name):
        name = honorific_pattern.sub("", name).strip()
    return name


def slugify(value: str) -> str:
    slug = re.sub(r"[^a-z0-9]+", "_", value.lower()).strip("_")
    return re.sub(r"_+", "_", slug)


def flat_key(building: str, flat_no: Any) -> str:
    number = int(float(flat_no))
    return f"{slugify(building)}_{number:03d}"


def flat_number(building: str, flat_no: Any) -> str:
    number = int(float(flat_no))
    return f"{building}-{number:03d}"


def safe_formula_number(formula: str) -> float | None:
    expr = formula.strip()
    if expr.startswith("="):
        expr = expr[1:]
    if expr.startswith("+"):
        expr = expr[1:]
    if not re.fullmatch(r"[0-9+\-*/(). ]+", expr):
        return None

    tree = ast.parse(expr, mode="eval")
    allowed = (
        ast.Expression,
        ast.BinOp,
        ast.UnaryOp,
        ast.Add,
        ast.Sub,
        ast.Mult,
        ast.Div,
        ast.USub,
        ast.UAdd,
        ast.Constant,
    )
    if not all(isinstance(node, allowed) for node in ast.walk(tree)):
        return None
    result = eval(compile(tree, "<formula>", "eval"), {"__builtins__": {}}, {})
    return float(result)


def cell_number(value: Any) -> float | None:
    if value is None or value == "":
        return None
    if isinstance(value, (int, float)):
        return float(value)
    if isinstance(value, str) and value.startswith("="):
        return safe_formula_number(value)
    try:
        return float(str(value).strip())
    except ValueError:
        return None


def source_rows(workbook: openpyxl.Workbook) -> list[dict[str, Any]]:
    ws = workbook["Total List"]
    rows: list[dict[str, Any]] = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        building, flat_no, owner, owner_phone, rental, tenant, tenant_phone, *_ = row
        if not building or not flat_no or not owner:
            continue
        key = flat_key(str(building), flat_no)
        owner_name = clean_person_name(owner)
        tenant_name = clean_person_name(tenant) if tenant else None
        rows.append(
            {
                "building": str(building).strip(),
                "original_flat_no": int(float(flat_no)),
                "flat_number": flat_number(str(building).strip(), flat_no),
                "owner_name": owner_name,
                "owner_phone": clean_phone(owner_phone),
                "is_rental": str(rental or "").strip().lower() == "yes",
                "tenant_name": tenant_name,
                "tenant_phone": clean_phone(tenant_phone),
                "slug": key,
                "email": f"gps.{key}@{GENERATED_DOMAIN}",
            }
        )
    return rows


def month_columns(ws: openpyxl.worksheet.worksheet.Worksheet) -> dict[int, str]:
    year_starts: list[tuple[int, int]] = []
    for col in range(1, ws.max_column + 1):
        value = ws.cell(6, col).value
        if isinstance(value, (int, float)) and 1900 <= int(value) <= 2100:
            year_starts.append((col, int(value)))

    mapping: dict[int, str] = {}
    for col in range(1, 108):
        year = None
        for start_col, start_year in year_starts:
            if start_col <= col:
                year = start_year
        label = ws.cell(7, col).value
        month = MONTH_NUMBERS.get(str(label or "").strip().upper())
        if year and month and year >= 2019:
            mapping[col] = f"{year}-{month:02d}"
    return mapping


def maintenance_rows(workbook: openpyxl.Workbook, residents: list[dict[str, Any]]) -> list[dict[str, Any]]:
    resident_by_slug = {item["slug"] for item in residents}
    ws = workbook["Monthwise maintenance"]
    columns = month_columns(ws)
    bills: list[dict[str, Any]] = []

    for row_idx in range(8, 86):
        sr_no = ws.cell(row_idx, 1).value
        building = ws.cell(row_idx, 2).value
        flat_no = ws.cell(row_idx, 4).value
        if not isinstance(sr_no, (int, float)) or not building or not flat_no:
            continue
        slug = flat_key(str(building).strip(), flat_no)
        if slug not in resident_by_slug:
            continue
        for col, period in columns.items():
            value = cell_number(ws.cell(row_idx, col).value)
            if value is None or value <= 0:
                continue
            bills.append(
                {
                    "slug": slug,
                    "period": period,
                    "amount": round(value, 2),
                    "source_cell": f"{get_column_letter(col)}{row_idx}",
                }
            )
    return bills


def render_values(name: str, columns: list[str], rows: list[list[Any]]) -> str:
    rendered = [f"{name}({', '.join(columns)}) AS (", "  VALUES"]
    for index, row in enumerate(rows):
        comma = "," if index < len(rows) - 1 else ""
        rendered.append("    (" + ", ".join(sql(value) for value in row) + ")" + comma)
    rendered.append(")")
    return "\n".join(rendered)


def build_markdown(residents: list[dict[str, Any]], bills: list[dict[str, Any]]) -> str:
    resident_values = render_values(
        "resident_source",
        [
            "building",
            "original_flat_no",
            "flat_number",
            "owner_name",
            "owner_phone",
            "is_rental",
            "tenant_name",
            "tenant_phone",
            "slug",
            "generated_email",
        ],
        [
            [
                row["building"],
                row["original_flat_no"],
                row["flat_number"],
                row["owner_name"],
                row["owner_phone"],
                row["is_rental"],
                row["tenant_name"],
                row["tenant_phone"],
                row["slug"],
                row["email"],
            ]
            for row in residents
        ],
    )

    bill_values = render_values(
        "bill_source",
        ["slug", "period", "amount", "source_cell"],
        [[row["slug"], row["period"], row["amount"], row["source_cell"]] for row in bills],
    )

    generated_at = dt.datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    years = sorted({row["period"][:4] for row in bills})
    period_range = f"{min(row['period'] for row in bills)} to {max(row['period'] for row in bills)}"

    return f"""# Godavari Puram SQL Import

Generated from `Dataset/GPS-Maintenance FY 2019 to 2026.xlsx` on {generated_at}.

## What This Imports

- 1 society: `{SOCIETY_NAME}`
- 1 active chairman/admin login
- {len(residents)} flats/units from `Total List`
- {len(residents)} owner `Person`, `User`, `SocietyMembership`, and `UnitOccupancy` records
- {len(bills)} historical maintenance entries from `Monthwise maintenance`, covering {period_range} ({", ".join(years)})

## Important Safety Notes

- The Excel file does not contain resident email addresses, so this SQL generates placeholder emails like `gps.nivrutti_a_001@godavaripuram.local`.
- All generated users get the temporary password `{TEMP_PASSWORD}`. Ask residents to change it after first login.
- The workbook repeats flat numbers across buildings. The database requires `Flat.flatNumber` to be unique per society, so this SQL imports flat numbers as `Building-FlatNo`, for example `Nivrutti-A-001`, and stores the building in `wing`.
- Historical maintenance cells appear to represent collection entries, not clean invoice definitions. This SQL imports non-empty positive cells as paid `MaintenanceBill` records and keeps the Excel source cell in `receiptNote`.
- Tenant names are mostly blank even when `Rental = Yes`, so this import creates owner accounts only. Rental flats are marked through legacy/unit occupancy fields but no tenant user is created without tenant data.
- Run this against a backup or staging database first. The SQL is idempotent for the deterministic IDs it creates, but it can still conflict if another society already uses join code `{JOIN_CODE}` or if generated emails were used manually.

## Login Details Created

- Chairman email: `chairman.godavari@godavaripuram.local`
- Resident email pattern: `gps.<building_slug>_<flat_no_3_digits>@godavaripuram.local`
- Temporary password for all generated users: `{TEMP_PASSWORD}`
- Society join code: `{JOIN_CODE}`

## SQL

```sql
BEGIN;

INSERT INTO "Society" (
  "id", "name", "joinCode", "address", "city", "pincode", "totalFlats",
  "maintenanceAmt", "dueDayOfMonth", "lateFee", "planTier", "openingBalance",
  "createdAt"
)
VALUES (
  {sql(SOCIETY_ID)}, {sql(SOCIETY_NAME)}, {sql(JOIN_CODE)},
  'TO_BE_UPDATED_FROM_REGISTRATION_RECORDS', 'TO_BE_UPDATED', '000000',
  {len(residents)}, 0, 10, 0, 'starter', 0, CURRENT_TIMESTAMP
)
ON CONFLICT ("id") DO UPDATE SET
  "name" = EXCLUDED."name",
  "joinCode" = EXCLUDED."joinCode",
  "totalFlats" = EXCLUDED."totalFlats";

INSERT INTO "User" (
  "id", "name", "email", "password", "phone", "role", "isActive",
  "societyId", "createdAt"
)
VALUES (
  {sql(ADMIN_USER_ID)}, 'Godavari Puram Chairman', 'chairman.godavari@godavaripuram.local',
  {sql(PASSWORD_HASH)}, NULL, 'chairman', TRUE, {sql(SOCIETY_ID)}, CURRENT_TIMESTAMP
)
ON CONFLICT ("id") DO UPDATE SET
  "name" = EXCLUDED."name",
  "email" = EXCLUDED."email",
  "password" = EXCLUDED."password",
  "phone" = EXCLUDED."phone",
  "role" = EXCLUDED."role",
  "isActive" = TRUE,
  "societyId" = EXCLUDED."societyId";

INSERT INTO "SocietyMembership" (
  "id", "userId", "societyId", "productRole", "permissionRole", "status",
  "approvedByUserId", "activatedAt", "createdAt", "updatedAt"
)
VALUES (
  {sql(ADMIN_MEMBERSHIP_ID)}, {sql(ADMIN_USER_ID)}, {sql(SOCIETY_ID)},
  'chairman', 'society_admin', 'active', {sql(ADMIN_USER_ID)},
  CURRENT_TIMESTAMP, CURRENT_TIMESTAMP, CURRENT_TIMESTAMP
)
ON CONFLICT ("userId", "societyId") DO UPDATE SET
  "productRole" = EXCLUDED."productRole",
  "permissionRole" = EXCLUDED."permissionRole",
  "status" = 'active',
  "approvedByUserId" = EXCLUDED."approvedByUserId",
  "activatedAt" = COALESCE("SocietyMembership"."activatedAt", CURRENT_TIMESTAMP),
  "updatedAt" = CURRENT_TIMESTAMP;

WITH
{resident_values}
INSERT INTO "Flat" (
  "id", "societyId", "flatNumber", "wing", "floor", "ownerName", "tenantName",
  "contact", "email", "flatType", "currentOccupant", "isActive", "createdAt", "updatedAt"
)
SELECT
  'flat_gps_' || slug,
  {sql(SOCIETY_ID)},
  flat_number,
  building,
  NULL,
  owner_name,
  tenant_name,
  owner_phone,
  generated_email,
  '2bhk',
  CASE WHEN is_rental THEN 'tenant' ELSE 'owner' END,
  TRUE,
  CURRENT_TIMESTAMP,
  CURRENT_TIMESTAMP
FROM resident_source
ON CONFLICT ("societyId", "flatNumber") DO UPDATE SET
  "wing" = EXCLUDED."wing",
  "ownerName" = EXCLUDED."ownerName",
  "tenantName" = EXCLUDED."tenantName",
  "contact" = EXCLUDED."contact",
  "email" = EXCLUDED."email",
  "currentOccupant" = EXCLUDED."currentOccupant",
  "isActive" = TRUE,
  "updatedAt" = CURRENT_TIMESTAMP;

WITH
{resident_values}
INSERT INTO "Unit" (
  "id", "societyId", "legacyFlatId", "wing", "flatNumber", "floor", "unitType",
  "usageType", "occupancyStatus", "billingStatus", "isActive", "createdAt", "updatedAt"
)
SELECT
  'unit_gps_' || slug,
  {sql(SOCIETY_ID)},
  'flat_gps_' || slug,
  building,
  flat_number,
  NULL,
  '2BHK',
  'RESIDENTIAL',
  CASE WHEN is_rental THEN 'TENANT_OCCUPIED' ELSE 'OWNER_OCCUPIED' END,
  'ACTIVE',
  TRUE,
  CURRENT_TIMESTAMP,
  CURRENT_TIMESTAMP
FROM resident_source
ON CONFLICT ("legacyFlatId") DO UPDATE SET
  "wing" = EXCLUDED."wing",
  "flatNumber" = EXCLUDED."flatNumber",
  "occupancyStatus" = EXCLUDED."occupancyStatus",
  "billingStatus" = 'ACTIVE',
  "isActive" = TRUE,
  "updatedAt" = CURRENT_TIMESTAMP;

WITH
{resident_values}
INSERT INTO "Person" (
  "id", "societyId", "name", "phone", "createdAt", "updatedAt"
)
SELECT
  'person_gps_' || slug || '_owner',
  {sql(SOCIETY_ID)},
  owner_name,
  owner_phone,
  CURRENT_TIMESTAMP,
  CURRENT_TIMESTAMP
FROM resident_source
ON CONFLICT ("id") DO UPDATE SET
  "name" = EXCLUDED."name",
  "phone" = EXCLUDED."phone",
  "updatedAt" = CURRENT_TIMESTAMP;

WITH
{resident_values}
INSERT INTO "User" (
  "id", "name", "email", "password", "phone", "role", "isActive",
  "societyId", "flatId", "personId", "showPhoneInDirectory", "showEmailInDirectory", "createdAt"
)
SELECT
  'user_gps_' || slug || '_owner',
  owner_name,
  generated_email,
  {sql(PASSWORD_HASH)},
  owner_phone,
  'member',
  TRUE,
  {sql(SOCIETY_ID)},
  'flat_gps_' || slug,
  'person_gps_' || slug || '_owner',
  TRUE,
  FALSE,
  CURRENT_TIMESTAMP
FROM resident_source
ON CONFLICT ("id") DO UPDATE SET
  "name" = EXCLUDED."name",
  "email" = EXCLUDED."email",
  "password" = EXCLUDED."password",
  "phone" = EXCLUDED."phone",
  "role" = 'member',
  "isActive" = TRUE,
  "societyId" = EXCLUDED."societyId",
  "flatId" = EXCLUDED."flatId",
  "personId" = EXCLUDED."personId";

WITH
{resident_values}
INSERT INTO "SocietyMembership" (
  "id", "userId", "societyId", "productRole", "permissionRole", "status",
  "flatId", "personId", "approvedByUserId", "activatedAt", "createdAt", "updatedAt"
)
SELECT
  'membership_gps_' || slug || '_owner',
  'user_gps_' || slug || '_owner',
  {sql(SOCIETY_ID)},
  'member',
  'member',
  'active',
  'flat_gps_' || slug,
  'person_gps_' || slug || '_owner',
  {sql(ADMIN_USER_ID)},
  CURRENT_TIMESTAMP,
  CURRENT_TIMESTAMP,
  CURRENT_TIMESTAMP
FROM resident_source
ON CONFLICT ("userId", "societyId") DO UPDATE SET
  "productRole" = EXCLUDED."productRole",
  "permissionRole" = EXCLUDED."permissionRole",
  "status" = 'active',
  "flatId" = EXCLUDED."flatId",
  "personId" = EXCLUDED."personId",
  "approvedByUserId" = EXCLUDED."approvedByUserId",
  "activatedAt" = COALESCE("SocietyMembership"."activatedAt", CURRENT_TIMESTAMP),
  "updatedAt" = CURRENT_TIMESTAMP;

WITH
{resident_values}
INSERT INTO "UnitOccupancy" (
  "id", "societyId", "unitId", "personId", "relationshipType", "occupancyStatus",
  "billingResponsibility", "vehicleCount", "isPrimaryOccupant", "isActive",
  "createdAt", "updatedAt"
)
SELECT
  'occupancy_gps_' || slug || '_owner',
  {sql(SOCIETY_ID)},
  'unit_gps_' || slug,
  'person_gps_' || slug || '_owner',
  'OWNER',
  'ACTIVE',
  'OWNER',
  0,
  TRUE,
  TRUE,
  CURRENT_TIMESTAMP,
  CURRENT_TIMESTAMP
FROM resident_source
ON CONFLICT ("id") DO UPDATE SET
  "occupancyStatus" = 'ACTIVE',
  "billingResponsibility" = 'OWNER',
  "isPrimaryOccupant" = TRUE,
  "isActive" = TRUE,
  "updatedAt" = CURRENT_TIMESTAMP;

WITH
{bill_values}
INSERT INTO "MaintenanceBill" (
  "id", "flatId", "societyId", "amount", "billType", "billingCycle", "description",
  "lateFee", "gstAmount", "totalAmount", "period", "dueDate", "status",
  "paidAt", "paidAmount", "receiptNote", "createdAt", "updatedAt"
)
SELECT
  'bill_gps_' || slug || '_' || replace(period, '-', '_'),
  'flat_gps_' || slug,
  {sql(SOCIETY_ID)},
  amount,
  'maintenance',
  'monthly',
  'Historical maintenance workbook entry imported from GPS-Maintenance FY 2019 to 2026.xlsx',
  0,
  0,
  amount,
  period,
  to_date(period || '-10', 'YYYY-MM-DD'),
  'paid',
  (
    date_trunc('month', to_date(period || '-01', 'YYYY-MM-DD')::timestamp)
    + interval '1 month - 1 day'
  )::timestamp,
  amount,
  'Source sheet Monthwise maintenance cell ' || source_cell,
  CURRENT_TIMESTAMP,
  CURRENT_TIMESTAMP
FROM bill_source
ON CONFLICT ("id") DO UPDATE SET
  "amount" = EXCLUDED."amount",
  "totalAmount" = EXCLUDED."totalAmount",
  "status" = EXCLUDED."status",
  "paidAt" = EXCLUDED."paidAt",
  "paidAmount" = EXCLUDED."paidAmount",
  "receiptNote" = EXCLUDED."receiptNote",
  "updatedAt" = CURRENT_TIMESTAMP;

INSERT INTO "ActivityLog" (
  "id", "societyId", "userId", "userName", "action", "module",
  "targetId", "targetLabel", "details", "createdAt"
)
VALUES (
  'activity_gps_import_2026_06_29',
  {sql(SOCIETY_ID)},
  {sql(ADMIN_USER_ID)},
  'Godavari Puram Chairman',
  'created',
  'settings',
  {sql(SOCIETY_ID)},
  {sql(SOCIETY_NAME)},
  {sql('Imported society, residents, active memberships, unit occupancies, and historical maintenance entries from workbook.')},
  CURRENT_TIMESTAMP
)
ON CONFLICT ("id") DO NOTHING;

COMMIT;
```

## Optional Pending-Approval Variant

If you want residents to appear in the membership approval queue instead of being immediately able to log in:

```sql
UPDATE "SocietyMembership"
SET "status" = 'pending', "activatedAt" = NULL, "approvedByUserId" = NULL, "updatedAt" = CURRENT_TIMESTAMP
WHERE "societyId" = {sql(SOCIETY_ID)}
  AND "productRole" IN ('member', 'tenant');

UPDATE "UnitOccupancy"
SET "occupancyStatus" = 'PENDING', "updatedAt" = CURRENT_TIMESTAMP
WHERE "societyId" = {sql(SOCIETY_ID)}
  AND "relationshipType" = 'OWNER';
```

## Post-Import Checks

```sql
SELECT COUNT(*) AS flats FROM "Flat" WHERE "societyId" = {sql(SOCIETY_ID)};
SELECT COUNT(*) AS users FROM "User" WHERE "societyId" = {sql(SOCIETY_ID)};
SELECT COUNT(*) AS active_memberships FROM "SocietyMembership" WHERE "societyId" = {sql(SOCIETY_ID)} AND "status" = 'active';
SELECT COUNT(*) AS active_occupancies FROM "UnitOccupancy" WHERE "societyId" = {sql(SOCIETY_ID)} AND "occupancyStatus" = 'ACTIVE';
SELECT COUNT(*) AS maintenance_bills FROM "MaintenanceBill" WHERE "societyId" = {sql(SOCIETY_ID)};
SELECT "period", COUNT(*) AS bills, SUM("paidAmount") AS collected
FROM "MaintenanceBill"
WHERE "societyId" = {sql(SOCIETY_ID)}
GROUP BY "period"
ORDER BY "period";
```
"""


def main() -> None:
    workbook = openpyxl.load_workbook(WORKBOOK, data_only=False, read_only=False)
    residents = source_rows(workbook)
    bills = maintenance_rows(workbook, residents)
    OUTPUT.parent.mkdir(parents=True, exist_ok=True)
    OUTPUT.write_text(build_markdown(residents, bills), encoding="utf-8")
    print(f"Wrote {OUTPUT}")
    print(f"Residents: {len(residents)}")
    print(f"Maintenance bills: {len(bills)}")


if __name__ == "__main__":
    main()
